"""Read one rectangular range out of an existing ``.xlsx`` and into the CV model.

Some content lives in a spreadsheet and nowhere else -- an invoice's line items,
kept as a workbook because that is what accounting actually sends around.
Retyping it as Markdown would lose both the grid (dates, currency, a totals row)
and the single source of truth, so instead the range is imported: a rectangle of
cells, named by its corners, converted to the same
:class:`~cv_generator.models.RichBlock` tree a ``.docx`` import produces (see
:mod:`cv_generator.docx_import`), and rendered by the normal backends.

There is exactly one section per import, and it is always one
:class:`~cv_generator.models.RichTable` -- a cell rectangle has no headings to
split it at, the same reason a ``.docx`` import is one section.

What "keeping the formatting" means here:

* **Carried over** -- bold, italic, underline, strikethrough, font size, whether
  any cell in the range is ruled, and the sheet's own column widths.
* **Left to the CV's theme** -- font family and everything about page layout, the
  same split the ``.docx`` import makes and for the same reason: the imported
  range has to sit in a document rendered by this project's themes, not carry a
  second document's design into it.
* **Always centered, and sized to its own content rather than the page.** Unlike
  a ``.docx`` import, which is meant to fill the page the way it did in the
  source document, a spreadsheet range reads more like a figure dropped into the
  document -- ``RichTable.centered`` is what tells both backends not to stretch
  it. ``column_widths`` still carries the sheet's own column proportions, but a
  centered table ignores them in favour of sizing each column to its content
  (see ``word.py::_add_imported_table`` and ``.cv-block-table--centered``).
* **Cell values are read, not formulas.** ``data_only`` reading returns the
  value Excel last calculated and cached into the file; a formula cell in a
  workbook that has never been opened in Excel since editing has nothing cached
  and reads as empty. That is a property of ``.xlsx`` itself, not something this
  module works around.
* **Number formatting is approximated, not interpreted.** Excel's format-code
  language is a small program in itself -- locales, conditional sections, dozens
  of tokens -- and this module does not implement it. It recognises just enough
  for a CV-adjacent spreadsheet: a date or datetime becomes ``DD.MM.YYYY``, a
  format naming a currency symbol becomes a two-decimal, thousands-grouped
  amount with that symbol, and anything else falls back to a plain rendering of
  the value.

Unlike ``begin``/``end``, which name headlines and stop *before* the second one,
the four corners a recipe gives here (``col-start``, ``col-end``, ``row-start``,
``row-end``) are Excel's own addressing and are all **inclusive** -- a range from
``C`` to ``G`` reads column G, not up to it. Excel-style corners get Excel's own
convention rather than one built for headings.
"""

from __future__ import annotations

import datetime
import zipfile
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from cv_generator.errors import CVParseError
from cv_generator.models import RichBlock, RichCell, RichParagraph, RichRun, RichTable

# A number format naming one of these is read as currency: the value is rounded
# to two decimals, grouped by thousands the German way (this project's own
# sample documents are all German), and the symbol is appended. Not every
# currency Excel knows -- just the ones plausible in a CV-adjacent spreadsheet.
_CURRENCY_SYMBOLS = ("€", "$", "£")


def load_section(
    path: Path, *, col_start: str, col_end: str, row_start: int, row_end: int
) -> list[RichBlock]:
    """Import the rectangle from ``col_start``/``row_start`` to ``col_end``/``row_end``.

    Args:
        col_start: The first column, as a letter (``"C"``). Matched
            case-insensitively, the way Excel itself would resolve it.
        col_end: The last column, inclusive.
        row_start: The first row, 1-based.
        row_end: The last row, inclusive.

    Returns:
        A single-element list holding one :class:`RichTable` -- always one,
        since a cell rectangle has no headings to split it at.

    Raises:
        CVParseError: if the file cannot be read as a workbook, has no active
            sheet, or the rectangle is empty or runs backwards.
    """
    workbook = _open(path)
    sheet = workbook.active
    if sheet is None:
        raise CVParseError(f"{path}: the workbook has no active sheet")

    first_col, last_col = _column_range(path, col_start, col_end)
    if row_start < 1:
        raise CVParseError(f"{path}: 'row-start' must be at least 1, got {row_start}")
    if row_start > row_end:
        raise CVParseError(f"{path}: 'row-start' {row_start} comes after 'row-end' {row_end}")

    grid = list(
        sheet.iter_rows(min_row=row_start, max_row=row_end, min_col=first_col, max_col=last_col)
    )
    table = RichTable(
        rows=[[_cell(cell) for cell in row] for row in grid],
        bordered=any(_is_bordered(cell) for row in grid for cell in row),
        column_widths=_column_widths(sheet, first_col, last_col),
        # A spreadsheet range is not meant to fill the page the way an imported
        # Word section is; it reads more like a figure set into the document, so
        # it is sized to its own content and centered rather than stretched.
        centered=True,
    )
    return [table]


def _open(path: Path) -> Workbook:
    try:
        # data_only: a formula cell's last calculated result, not its formula --
        # this project renders a value, not a spreadsheet engine.
        return load_workbook(str(path), data_only=True)
    except (InvalidFileException, zipfile.BadZipFile, KeyError, ValueError, OSError) as exc:
        raise CVParseError(f"cannot read {path} as an Excel workbook: {exc}") from exc


def _column_range(path: Path, col_start: str, col_end: str) -> tuple[int, int]:
    try:
        first = column_index_from_string(col_start.strip().upper())
        last = column_index_from_string(col_end.strip().upper())
    except ValueError as exc:
        raise CVParseError(f"{path}: not a column letter: {exc}") from exc
    if first > last:
        raise CVParseError(f"{path}: 'col-start' {col_start!r} comes after 'col-end' {col_end!r}")
    return first, last


# -- cells ------------------------------------------------------------------


def _cell(cell: Cell) -> RichCell:
    """One cell, as the one paragraph it is -- Excel has no sub-cell structure."""
    text = _display_value(cell)
    runs = (
        []
        if not text
        else [
            RichRun(
                text=text,
                bold=bool(cell.font.bold),
                italic=bool(cell.font.italic),
                underline=bool(cell.font.underline),
                strike=bool(cell.font.strikethrough),
                size_pt=cell.font.size,
            )
        ]
    )
    return RichCell(paragraphs=[RichParagraph(runs=runs)])


def _is_bordered(cell: Cell) -> bool:
    border = cell.border
    return any(
        side is not None and side.style is not None
        for side in (border.left, border.right, border.top, border.bottom)
    )


def _column_widths(sheet: Worksheet, first_col: int, last_col: int) -> list[float]:
    widths = [
        sheet.column_dimensions[get_column_letter(index)].width
        for index in range(first_col, last_col + 1)
    ]
    total = sum(width for width in widths if width is not None)
    if not total or any(width is None for width in widths):
        return []
    return [round(float(width) / float(total), 4) for width in widths if width is not None]


# -- values -------------------------------------------------------------


def _display_value(cell: Cell) -> str:
    value = cell.value
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, datetime.time):
        return value.strftime("%H:%M")
    if isinstance(value, datetime.date):
        # Covers `datetime.datetime` too (a `date` subclass); the time of day, if
        # any, is dropped -- this project's own dates are day-level.
        return value.strftime("%d.%m.%Y")
    if isinstance(value, (int, float)):
        return _format_number(float(value), cell.number_format)
    return str(value)


def _format_number(value: float, number_format: str) -> str:
    symbol = next((s for s in _CURRENCY_SYMBOLS if s in number_format), None)
    if symbol is not None:
        grouped = f"{value:,.2f}"  # "1,275.00"
        grouped = grouped.replace(",", "\x00").replace(".", ",").replace("\x00", ".")
        return f"{grouped} {symbol}"
    if value.is_integer():
        return str(int(value))
    return f"{value:g}"


__all__ = ["load_section"]
